"""
Read-only helper tools for onboarding a new Vehicle into the Reference Workbook (ADR-013).

This script never writes to the workbook. It exists to remove two specific
sources of manual friction found while onboarding vehicle #6 (BYD Atto 3):

1. Computing the next available ID for a worksheet by hand (error-prone,
   and the "next row count + 1" shortcut is wrong whenever a sheet has a
   historical gap - e.g. 10_Scoring has 120 data rows but its max ID is
   SCORE_000121).
2. Sanity-checking the expected OverallScore/CoveragePercent for a
   Configuration without needing LibreOffice/Excel to recalculate the
   workbook's live formulas (10_Scoring.RawScore/WeightedScore and
   12_OverallScores.OverallScore/CoveragePercent).

The scoring formulas reimplemented here are the workbook's own, defined in
ADR-005 (Scoring Mechanics) and documented in docs/03_scoring-model.md
Stage 5-6. If a future *Methodology*-type Framework Version bump (ADR-010)
changes those formulas, this script must be updated to match - it is a
sanity check against the workbook, never a replacement for it.
"""

import argparse
import re
import sys
from pathlib import Path

import openpyxl

DEFAULT_WORKBOOK = Path(__file__).resolve().parent.parent / "data" / "EV_Decision_Framework.xlsx"

# Worksheet -> (prefix, digit width). Mirrors framework/architecture/workbook-schema.md's
# Naming Conventions table (ADR-013). Kept here as a static copy for a minimal,
# dependency-light script - workbook-schema.md remains the authoritative source
# and this table must be kept in sync with it manually.
PREFIX_TABLE = {
    "04_Technical": ("TECH_", 6),
    "05_Equipment": ("EQ_", 6),
    "07_Reviews": ("REV_", 6),
    "08_Evidence": ("EV_", 6),
    "10_Scoring": ("SCORE_", 6),
    "12_OverallScores": ("OVSC_", 6),
    "14_HardRequirementResults": ("HRR_", 6),
}

FREE_TEXT_SHEETS = {"09_Sources", "02_Vehicles", "03_Configurations"}

# The 7 Configurations that currently receive full Evidence/Review/Score/OverallScore
# treatment (see workbook-schema.md's Flagship Configuration Scope, ADR-013).
FLAGSHIP_CONFIGURATION_IDS = [
    "TESLA_MODEL_3_LONG_RANGE_RWD",
    "RENAULT_4_TECHNO",
    "SKODA_ELROQ_SELECTION_60",
    "VOLVO_EX30_P5_LONG_RANGE",
    "KIA_EV2_FWD_LONG_RANGE_GT_LINE",
    "BYD_ATTO3_DESIGN",
    "SKODA_EPIQ_SELECTION_55_LAUNCH_EDITION",
]

# sheet -> (VehicleID column name or None, ConfigurationID column name).
# Mirrors 15_Dashboard's Coverage Grid formulas - kept in sync manually.
COVERAGE_SHEETS = {
    "04_Technical": ("VehicleID", "ConfigurationID"),
    "05_Equipment": (None, "ConfigurationID"),
    "08_Evidence": ("VehicleID", "ConfigurationID"),
    "07_Reviews": ("VehicleID", "ConfigurationID"),
    "10_Scoring": (None, "ConfigurationID"),
}


def load(path):
    return openpyxl.load_workbook(path, data_only=False)


def cmd_next_id(args):
    wb = load(args.workbook)

    if args.sheet == "11_DecisionLog":
        if not args.category:
            sys.exit("11_DecisionLog requires --category (e.g. --category ADR)")
        ws = wb[args.sheet]
        pattern = re.compile(rf"^DEC_{re.escape(args.category)}_(\d{{3}})$")
        max_n = 0
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            cell = row[0]
            if cell:
                m = pattern.match(str(cell))
                if m:
                    max_n = max(max_n, int(m.group(1)))
        print(f"DEC_{args.category}_{max_n + 1:03d}")
        return

    if args.sheet in FREE_TEXT_SHEETS:
        sys.exit(
            f"{args.sheet} uses free-text, name-based identifiers - pick one manually "
            f"(see workbook-schema.md's Naming Conventions table)."
        )

    if args.sheet not in PREFIX_TABLE:
        sys.exit(f"Unknown or unsupported worksheet: {args.sheet}")

    prefix, digits = PREFIX_TABLE[args.sheet]
    ws = wb[args.sheet]
    pattern = re.compile(rf"^{re.escape(prefix)}(\d{{{digits}}})$")
    max_n = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        cell = row[0]
        if cell:
            m = pattern.match(str(cell))
            if m:
                max_n = max(max_n, int(m.group(1)))
    print(f"{prefix}{max_n + 1:0{digits}d}")


def _criteria_weights(wb):
    """CriterionID -> Weight, restricted to WEIGHTED + Active criteria."""
    ws = wb["01_Criteria"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    weights = {}
    total_weight = 0.0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["Type"]] == "WEIGHTED" and str(row[idx["Active"]]).upper() == "TRUE":
            weight = float(row[idx["Weight"]])
            weights[row[idx["CriterionID"]]] = weight
            total_weight += weight
    return weights, total_weight


def _review_scores(wb):
    """ReviewID -> Score (1-5)."""
    ws = wb["07_Reviews"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    return {
        row[idx["ReviewID"]]: float(row[idx["Score"]])
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[idx["ReviewID"]]
    }


def cmd_score(args):
    wb = load(args.workbook)
    weights, total_weight = _criteria_weights(wb)
    review_scores = _review_scores(wb)

    ws = wb["10_Scoring"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    rows = [
        row
        for row in ws.iter_rows(min_row=2, values_only=True)
        if row[idx["ConfigurationID"]] == args.configuration_id
    ]
    if not rows:
        sys.exit(f"No 10_Scoring rows found for {args.configuration_id}")

    framework_versions = {str(row[idx["FrameworkVersion"]]) for row in rows}
    if args.framework_version:
        framework_version = args.framework_version
    elif len(framework_versions) == 1:
        framework_version = next(iter(framework_versions))
    else:
        sys.exit(
            f"{args.configuration_id} has rows across mixed FrameworkVersions "
            f"({sorted(framework_versions)}) - pass --framework-version explicitly."
        )

    print(f"ConfigurationID: {args.configuration_id}")
    print(f"FrameworkVersion: {framework_version}")
    print(f"{'CriterionID':<12}{'Weight':>8}{'Review.Score':>14}{'RawScore':>10}{'WeightedScore':>15}")

    overall_score = 0.0
    criteria_scored = 0
    for row in rows:
        if str(row[idx["FrameworkVersion"]]) != framework_version:
            continue
        criterion_id = row[idx["CriterionID"]]
        review_id = row[idx["ReviewID"]]
        weight = weights.get(criterion_id)
        score = review_scores.get(review_id)
        if weight is None or score is None:
            print(f"  WARNING: skipping {criterion_id} ({review_id}) - missing weight or review score")
            continue
        raw_score = round(score / 5 * 100, 2)
        weighted_score = round(raw_score * weight / 100, 2)
        overall_score += weighted_score
        criteria_scored += 1
        print(f"{criterion_id:<12}{weight:>8}{score:>14}{raw_score:>10}{weighted_score:>15}")

    criteria_total = len(weights)
    coverage_percent = round(
        sum(weights[r[idx["CriterionID"]]] for r in rows if r[idx["CriterionID"]] in weights)
        / total_weight
        * 100,
        2,
    ) if total_weight else 0.0

    print()
    print(f"OverallScore:     {round(overall_score, 2)}")
    print(f"CoveragePercent:  {coverage_percent}")
    print(f"CriteriaScored:   {criteria_scored}")
    print(f"CriteriaTotal:    {criteria_total}")
    print()
    print(
        "Note: this does not replicate 12_OverallScores.Notes' hardcoded override-flagging "
        "text (e.g. a Tesla-style 'HARD REQUIREMENT FAILED' string). Check that column "
        "manually for eligibility overrides."
    )


def _parent_vehicle_id(wb, configuration_id):
    ws = wb["03_Configurations"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["ConfigurationID"]] == configuration_id:
            return row[idx["VehicleID"]]
    return None


def _coverage_counts(wb, configuration_id, vehicle_id):
    """Independently recomputes 15_Dashboard's Coverage Grid formulas for one Configuration."""
    counts = {}
    for sheet_name, (vehicle_col, config_col) in COVERAGE_SHEETS.items():
        ws = wb[sheet_name]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(header)}
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if vehicle_col and row[idx[vehicle_col]] == vehicle_id:
                count += 1
            elif row[idx[config_col]] == configuration_id:
                count += 1
        counts[sheet_name] = count
    return counts


def cmd_coverage(args):
    wb = load(args.workbook)
    configuration_ids = args.configuration_id or FLAGSHIP_CONFIGURATION_IDS

    print(
        f"{'ConfigurationID':<32}{'Technical':>10}{'Equipment':>10}"
        f"{'Evidence':>10}{'Reviews':>10}{'Score':>8}"
    )
    for config_id in configuration_ids:
        vehicle_id = _parent_vehicle_id(wb, config_id)
        if vehicle_id is None:
            print(f"  WARNING: {config_id} not found in 03_Configurations - skipping")
            continue
        counts = _coverage_counts(wb, config_id, vehicle_id)
        print(
            f"{config_id:<32}"
            f"{counts['04_Technical']:>10}{counts['05_Equipment']:>10}"
            f"{counts['08_Evidence']:>10}{counts['07_Reviews']:>10}{counts['10_Scoring']:>8}"
        )

    print()
    print("Ranked Decision Summary (OverallScore descending):")
    print(f"{'Rank':<6}{'ConfigurationID':<32}{'OverallScore':>14}{'CoveragePercent':>17}")
    ranked = []
    for config_id in configuration_ids:
        weights, total_weight = _criteria_weights(wb)
        review_scores = _review_scores(wb)
        ws = wb["10_Scoring"]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        idx = {name: i for i, name in enumerate(header)}
        overall_score = 0.0
        scored_weight = 0.0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[idx["ConfigurationID"]] != config_id:
                continue
            criterion_id = row[idx["CriterionID"]]
            review_id = row[idx["ReviewID"]]
            weight = weights.get(criterion_id)
            score = review_scores.get(review_id)
            if weight is None or score is None:
                continue
            raw_score = round(score / 5 * 100, 2)
            overall_score += round(raw_score * weight / 100, 2)
            scored_weight += weight
        coverage_percent = round(scored_weight / total_weight * 100, 2) if total_weight else 0.0
        ranked.append((config_id, round(overall_score, 2), coverage_percent))

    for rank, (config_id, overall_score, coverage_percent) in enumerate(
        sorted(ranked, key=lambda t: t[1], reverse=True), start=1
    ):
        print(f"{rank:<6}{config_id:<32}{overall_score:>14}{coverage_percent:>17}")

    print()
    print(
        "Note: this recomputes what 15_Dashboard's formulas are expected to show. It does not "
        "confirm the chart or conditional-formatting colors actually render correctly - open "
        "the workbook in Excel to check those visually."
    )


def cmd_validate(args):
    wb_formulas = load(args.workbook)
    wb_values = openpyxl.load_workbook(args.workbook, data_only=True)

    ws = wb_values["12_OverallScores"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}

    live_row = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[idx["ConfigurationID"]] == args.configuration_id:
            live_row = row
            break

    if live_row is None:
        sys.exit(f"No 12_OverallScores row found for {args.configuration_id}")

    if live_row[idx["OverallScore"]] is None:
        print(
            "12_OverallScores has a row for this Configuration, but its formulas have never "
            "been recalculated (cached value is empty) - open the workbook in Excel/LibreOffice "
            "once, save, and re-run this command to compare. Falling back to `score` output only."
        )
        cmd_score(args)
        return

    print(f"Live workbook: OverallScore={live_row[idx['OverallScore']]}, "
          f"CoveragePercent={live_row[idx['CoveragePercent']]}")
    print()
    cmd_score(args)


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK), help="Path to the .xlsx workbook")
    subparsers = parser.add_subparsers(dest="command", required=True)

    p_next_id = subparsers.add_parser("next-id", help="Print the next available ID for a worksheet")
    p_next_id.add_argument("sheet", help="Worksheet name, e.g. 04_Technical")
    p_next_id.add_argument("--category", help="Required for 11_DecisionLog, e.g. ADR or DATA")
    p_next_id.set_defaults(func=cmd_next_id)

    p_score = subparsers.add_parser(
        "score", help="Recompute OverallScore/CoveragePercent for a Configuration without LibreOffice"
    )
    p_score.add_argument("configuration_id")
    p_score.add_argument("--framework-version", help="Override auto-detected FrameworkVersion")
    p_score.set_defaults(func=cmd_score)

    p_validate = subparsers.add_parser(
        "validate", help="Compare the dry-run score against the live 12_OverallScores row, if recalculated"
    )
    p_validate.add_argument("configuration_id")
    p_validate.add_argument("--framework-version", help="Override auto-detected FrameworkVersion")
    p_validate.set_defaults(func=cmd_validate)

    p_coverage = subparsers.add_parser(
        "coverage",
        help="Independently recompute what 15_Dashboard's Coverage Grid and Decision Summary should show",
    )
    p_coverage.add_argument(
        "--configuration-id",
        action="append",
        help="Repeatable. Defaults to the 6 flagship Configurations if omitted.",
    )
    p_coverage.set_defaults(func=cmd_coverage)

    args = parser.parse_args()
    args.func(args)


if __name__ == "__main__":
    main()
